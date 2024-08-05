#'''
# Generate part of the original author list
# The authors of the 10 most highly cited papers for the top 10 ERE journals

import requests
import pandas as pd
import os
from openpyxl import load_workbook

# Create session
session = requests.Session()


# Fetch authors' information
def get_authors_info(work_id):
    url = f"https://api.openalex.org/works/{work_id}"
    response = session.get(url)
    if response.status_code == 200:
        data = response.json()
        authors_info = []
        for author in data.get('authorships', []):
            author_name = author.get('author', {}).get('display_name', 'N/A')
            author_url = author.get('author', {}).get('id', 'N/A')
            if author_url.startswith("https://openalex.org/"):
                author_url = author_url
            authors_info.append({"author_name": author_name, "author_url": author_url})
        return authors_info
    else:
        print(f"Failed to fetch data for {work_id}: {response.status_code}")
        return []


# Process Excel files
def process_excel_file(input_file, output_file):
    # Load Excel files
    book = load_workbook(input_file)

    results = []

    for sheet_name in book.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        df = pd.read_excel(input_file, sheet_name=sheet_name)

        for index, row in df.iterrows():
            work_id = row['id'].split('/')[-1]  # Get work ID
            authors_info = get_authors_info(work_id)
            for author in authors_info:
                results.append({
                    'work_id': row['id'],
                    'author_name': author['author_name'],
                    'author_url': author['author_url']
                })

    # Save results as a new Excel file
    if results:
        df_results = pd.DataFrame(results)
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_results.to_excel(writer, index=False, sheet_name="Authors Info")

    print(f"Results saved to {output_file}")


# Main function
def main():
    input_file = 'the file that stores work id about the 10 most highly cited papers for the top 10 ERE journals'
    output_file = 'the file that stores the author lists'

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    process_excel_file(input_file, output_file)


if __name__ == "__main__":
    main()
#'''


#'''
# Rank top 30 authors in the field of
# environmental and resource economics
import requests
import pandas as pd
import os
import time
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Create session
session = requests.Session()

# Set file paths
input_file = "the file that stores the author lists"
output_folder = "the file folder that stores authors' publications in ERE fields"
output_file_top_30 = os.path.join(output_folder, "original_top_30_authors.xlsx")


# define the designated key terms
keywords = ["Environment", "Environmental", "Pollution", "Energy", "Climate", "Carbon", "Resource", "Resources"]

# Fetch the publications of authors and filter those in the ERE fields
def get_author_works(author_id):
    base_url = "https://api.openalex.org/works"
    params = {
        "filter": f"authorships.author.id:{author_id}",
        "per-page": 200,
        "cursor": "*"
    }

    works = []
    while True:
        try:
            response = session.get(base_url, params=params, headers={'Accept-Encoding': 'identity'})
            if response.status_code == 200:
                data = response.json()
                if 'results' in data:
                    works.extend(data['results'])
                next_cursor = data.get('meta', {}).get('next_cursor')
                print(f"Next cursor: {next_cursor}")
                if next_cursor:
                    params['cursor'] = next_cursor
                else:
                    break
            else:
                print(f"Failed to fetch data: {response.status_code}")
                print("Error response:", response.text)
                break
        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}")
            break


        time.sleep(1)

    return works

# filter publication works with designated terms
# in the API fields of 'title', 'keywords', 'topics', and 'concepts'
def filter_works_by_keywords(works, keywords):
    filtered_works = []
    for work in works:
        title = work.get('title', '')
        concepts = [concept['display_name'] for concept in work.get('concepts', []) if isinstance(concept['display_name'], str)]
        keywords_list = [kw for kw in work.get('keywords', []) if isinstance(kw, str)]
        topics = [topic['display_name'] for topic in work.get('topics', []) if isinstance(topic['display_name'], str)]

        combined_list = concepts + keywords_list + topics + [title]

        if any(keyword.lower() in item.lower() for keyword in keywords for item in combined_list if isinstance(item, str)):
            filtered_works.append(work)
    return filtered_works

# process Excel file
def process_excel_file(input_file, output_folder):
    # Load Excel file
    df = pd.read_excel(input_file)

    author_citations = []

    for index, row in df.iterrows():
        author_name = row['author_name']
        author_id = row['author_url'].split('/')[-1]

        print(f"Processing author: {author_name}")

        # Fetch author information
        works = get_author_works(author_id)

        # filter publication works with designated terms
        filtered_works = filter_works_by_keywords(works, keywords)

        # Count total citations
        total_citations = sum(work.get('cited_by_count', 0) for work in filtered_works)

        # Save each author's total citations
        author_citations.append({
            'author_name': author_name,
            'total_citations': total_citations,
            'filtered_works': filtered_works
        })

    # Rank total citations, get top 30 authors with the highest citations
    top_30_authors = sorted(author_citations, key=lambda x: x['total_citations'], reverse=True)[:30]

    # Save the results of the top 30 authors to a new Excel file
    with pd.ExcelWriter(output_file_top_30, engine='openpyxl') as writer:
        for author in top_30_authors:
            author_name = author['author_name']
            filtered_works = author['filtered_works']

            if filtered_works:
                df_filtered = pd.DataFrame(filtered_works)
                sheet_name = author_name[:31]  # Sheet names cannot exceed 31 characters
                df_filtered.to_excel(writer, index=False, sheet_name=sheet_name)

    print(f"Top 30 authors' results saved to {output_file_top_30}")

# Main function
def main():
    os.makedirs(output_folder, exist_ok=True)
    process_excel_file(input_file, output_folder)

if __name__ == "__main__":
    main()
#'''
# Then manually filter each author's publication in ERE fields:
# 1. To look at an author’s 20/25/30 most highly cited works in ERE fields
# 2. Exclude citation counts in the textbooks for undergrads and popular science
# 3. Exclude papers in science journals

